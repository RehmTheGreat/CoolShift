"""Excel and CSV data importer service."""

import os
import io
from datetime import datetime
from typing import BinaryIO, Dict, Any, List
import pandas as pd
import openpyxl
from sqlalchemy.orm import Session

from app.models import (
    ScenarioProfile,
    Appliance,
    EnergyAsset,
    IntervalInput,
    BaselineSchedule,
)

def parse_and_import_excel(file_content: bytes, db: Session) -> Dict[str, int]:
    """Parses CoolShift excel sheets and loads them into database.
    
    Returns a dictionary of counts of imported rows per model.
    """
    # Write to a temp file to read with openpyxl/pandas
    temp_filename = "temp_import.xlsx"
    with open(temp_filename, "wb") as f:
        f.write(file_content)
        
    counts = {
        "scenario_profiles": 0,
        "appliances": 0,
        "energy_assets": 0,
        "interval_inputs": 0,
        "baseline_schedules": 0,
    }
    
    try:
        xls = pd.ExcelFile(temp_filename)
        
        # 1. Scenario Profiles
        if "Scenario_Profiles" in xls.sheet_names:
            df = pd.read_excel(xls, "Scenario_Profiles")
            # Clean columns and rows
            df = df.dropna(subset=["scenario_id"])
            for _, row in df.iterrows():
                scenario_id = str(row["scenario_id"]).strip()
                # Check if exists, update or insert
                profile = db.query(ScenarioProfile).filter_by(scenario_id=scenario_id).first()
                if not profile:
                    profile = ScenarioProfile(scenario_id=scenario_id)
                    db.add(profile)
                
                profile.name = str(row["name"])
                profile.timezone = str(row.get("timezone", "Asia/Karachi"))
                profile.building_type = str(row["building_type"])
                profile.area_m2 = float(row["area_m2"])
                profile.room_count = int(row["room_count"])
                profile.max_occupancy = int(row["max_occupancy"])
                profile.insulation_level = str(row["insulation_level"])
                profile.sun_exposure = str(row["sun_exposure"])
                profile.comfort_min_c = float(row["comfort_min_c"])
                profile.comfort_max_c = float(row["comfort_max_c"])
                profile.vulnerable_occupants = bool(row["vulnerable_occupants"]) if pd.notna(row["vulnerable_occupants"]) else False
                profile.budget_pkr_per_day = float(row["budget_pkr_per_day"])
                profile.maximum_grid_demand_kw = float(row["maximum_grid_demand_kw"])
                counts["scenario_profiles"] += 1
            db.commit()
            
        # 2. Appliances
        if "Appliances" in xls.sheet_names:
            df = pd.read_excel(xls, "Appliances")
            df = df.dropna(subset=["scenario_id", "appliance_id"])
            # Clear existing appliances for the scenario_ids we see
            scenarios_seen = df["scenario_id"].dropna().unique()
            for s_id in scenarios_seen:
                db.query(Appliance).filter_by(scenario_id=str(s_id).strip()).delete()
                
            for _, row in df.iterrows():
                app = Appliance(
                    scenario_id=str(row["scenario_id"]).strip(),
                    appliance_id=str(row["appliance_id"]).strip(),
                    zone_id=str(row["zone_id"]).strip(),
                    appliance_type=str(row["appliance_type"]).strip(),
                    quantity=int(row["quantity"]),
                    rated_power_kw=float(row["rated_power_kw"]),
                    cooling_capacity_kw=float(row["cooling_capacity_kw"]) if pd.notna(row["cooling_capacity_kw"]) else 0.0,
                    efficiency_label=str(row["efficiency_label"]) if pd.notna(row["efficiency_label"]) else None,
                    min_runtime_minutes=int(row["min_runtime_minutes"]) if pd.notna(row["min_runtime_minutes"]) else 0,
                    min_setpoint_c=float(row["min_setpoint_c"]) if pd.notna(row["min_setpoint_c"]) else None,
                    max_setpoint_c=float(row["max_setpoint_c"]) if pd.notna(row["max_setpoint_c"]) else None,
                )
                db.add(app)
                counts["appliances"] += 1
            db.commit()
            
        # 3. Energy Assets
        if "Energy_Assets" in xls.sheet_names:
            df = pd.read_excel(xls, "Energy_Assets")
            df = df.dropna(subset=["scenario_id"])
            for _, row in df.iterrows():
                scenario_id = str(row["scenario_id"]).strip()
                asset = db.query(EnergyAsset).filter_by(scenario_id=scenario_id).first()
                if not asset:
                    asset = EnergyAsset(scenario_id=scenario_id)
                    db.add(asset)
                    
                asset.solar_capacity_kw = float(row["solar_capacity_kw"]) if pd.notna(row["solar_capacity_kw"]) else 0.0
                asset.solar_conversion_efficiency = float(row["solar_conversion_efficiency"]) if pd.notna(row["solar_conversion_efficiency"]) else 0.0
                asset.battery_capacity_kwh = float(row["battery_capacity_kwh"]) if pd.notna(row["battery_capacity_kwh"]) else 0.0
                asset.initial_soc_kwh = float(row["initial_soc_kwh"]) if pd.notna(row["initial_soc_kwh"]) else 0.0
                asset.minimum_reserve_kwh = float(row["minimum_reserve_kwh"]) if pd.notna(row["minimum_reserve_kwh"]) else 0.0
                asset.max_charge_kw = float(row["max_charge_kw"]) if pd.notna(row["max_charge_kw"]) else 0.0
                asset.max_discharge_kw = float(row["max_discharge_kw"]) if pd.notna(row["max_discharge_kw"]) else 0.0
                asset.charge_efficiency = float(row["charge_efficiency"]) if pd.notna(row["charge_efficiency"]) else 1.0
                asset.discharge_efficiency = float(row["discharge_efficiency"]) if pd.notna(row["discharge_efficiency"]) else 1.0
                counts["energy_assets"] += 1
            db.commit()
            
        # 4. Interval Inputs
        if "Interval_Inputs" in xls.sheet_names:
            df = pd.read_excel(xls, "Interval_Inputs")
            df = df.dropna(subset=["scenario_id", "timestamp_local"])
            
            scenarios_seen = df["scenario_id"].dropna().unique()
            for s_id in scenarios_seen:
                db.query(IntervalInput).filter_by(scenario_id=str(s_id).strip()).delete()
                
            intervals = []
            for _, row in df.iterrows():
                # Convert timestamp to python datetime object
                ts = row["timestamp_local"]
                ts_dt = ts.to_pydatetime() if hasattr(ts, "to_pydatetime") else pd.to_datetime(ts).to_pydatetime()
                    
                # Calculate heat index if missing or zero
                hi = float(row["heat_index_c"]) if pd.notna(row["heat_index_c"]) else float(row["temperature_c"])
                
                # Available solar calculation fallback
                solar_av = float(row["solar_available_kw"]) if pd.notna(row.get("solar_available_kw")) else 0.0
                
                # Carbon factor fallback
                carbon = float(row["grid_carbon_kgco2_per_kwh"]) if pd.notna(row.get("grid_carbon_kgco2_per_kwh")) else 0.46
                
                # Check grid available
                grid_av = bool(row["grid_available"]) if pd.notna(row["grid_available"]) else True
                
                # Tariff type
                t_type = str(row["tariff_type"]).strip().lower() if pd.notna(row["tariff_type"]) else "off_peak"
                
                # Missing flag
                missing = bool(row["source_missing_flag"]) if pd.notna(row.get("source_missing_flag")) else False
                
                interval = IntervalInput(
                    scenario_id=str(row["scenario_id"]).strip(),
                    timestamp_local=ts_dt,
                    interval_minutes=int(row["interval_minutes"]) if pd.notna(row["interval_minutes"]) else 15,
                    temperature_c=float(row["temperature_c"]),
                    relative_humidity_pct=float(row["relative_humidity_pct"]),
                    heat_index_c=hi,
                    solar_irradiance_w_m2=float(row["solar_irradiance_w_m2"]) if pd.notna(row["solar_irradiance_w_m2"]) else 0.0,
                    solar_available_kw=solar_av,
                    occupancy_count=int(row["occupancy_count"]) if pd.notna(row["occupancy_count"]) else 0,
                    grid_available=grid_av,
                    tariff_type=t_type,
                    tariff_pkr_per_kwh=float(row["tariff_pkr_per_kwh"]),
                    grid_carbon_kgco2_per_kwh=carbon,
                    non_cooling_load_kw=float(row["non_cooling_load_kw"]) if pd.notna(row["non_cooling_load_kw"]) else 0.0,
                    source_missing_flag=missing,
                )
                intervals.append(interval)
                counts["interval_inputs"] += 1
                
            db.bulk_save_objects(intervals)
            db.commit()
            
        # 5. Baseline Schedule
        if "Baseline_Schedule" in xls.sheet_names:
            df = pd.read_excel(xls, "Baseline_Schedule")
            df = df.dropna(subset=["scenario_id", "timestamp_local"])
            
            scenarios_seen = df["scenario_id"].dropna().unique()
            for s_id in scenarios_seen:
                db.query(BaselineSchedule).filter_by(scenario_id=str(s_id).strip()).delete()
                
            baselines = []
            for _, row in df.iterrows():
                ts = row["timestamp_local"]
                ts_dt = ts.to_pydatetime() if hasattr(ts, "to_pydatetime") else pd.to_datetime(ts).to_pydatetime()
                    
                ac_on = int(row["baseline_ac_units_on"]) if pd.notna(row["baseline_ac_units_on"]) else 0
                setpoint = float(row["baseline_ac_setpoint_c"]) if pd.notna(row["baseline_ac_setpoint_c"]) else None
                fan_on = int(row["baseline_fan_units_on"]) if pd.notna(row["baseline_fan_units_on"]) else 0
                other = float(row["baseline_other_cooling_kw"]) if pd.notna(row["baseline_other_cooling_kw"]) else 0.0
                rule = str(row["baseline_rule"]) if pd.notna(row["baseline_rule"]) else None
                
                baseline = BaselineSchedule(
                    scenario_id=str(row["scenario_id"]).strip(),
                    timestamp_local=ts_dt,
                    baseline_ac_units_on=ac_on,
                    baseline_ac_setpoint_c=setpoint,
                    baseline_fan_units_on=fan_on,
                    baseline_other_cooling_kw=other,
                    baseline_rule=rule,
                )
                baselines.append(baseline)
                counts["baseline_schedules"] += 1
                
            db.bulk_save_objects(baselines)
            db.commit()
            
    finally:
        if 'xls' in locals():
            try:
                xls.close()
            except Exception:
                pass
        if os.path.exists(temp_filename):
            os.remove(temp_filename)
            
    return counts


def import_excel_with_validation(file_source: str | bytes | BinaryIO, db: Session) -> Dict[str, Any]:
    """Reads CoolShift Excel workbook using openpyxl, performs strict validations, and imports data.
    
    Returns a structured dictionary:
        { "success": bool, "errors": List[str], "warnings": List[str], "imported_counts": Dict[str, int] }
    """
    errors = []
    warnings = []
    imported_counts = {
        "scenario_profiles": 0,
        "appliances": 0,
        "energy_assets": 0,
        "interval_inputs": 0,
        "baseline_schedules": 0,
    }
    
    # Load workbook
    wb = None
    try:
        if isinstance(file_source, bytes):
            wb = openpyxl.load_workbook(io.BytesIO(file_source), data_only=True)
        elif isinstance(file_source, str):
            wb = openpyxl.load_workbook(file_source, data_only=True)
        else:
            wb = openpyxl.load_workbook(file_source, data_only=True)
    except Exception as e:
        return {
            "success": False,
            "errors": [f"Failed to open Excel workbook: {str(e)}"],
            "warnings": [],
            "imported_counts": imported_counts
        }
        
    def find_sheet_header(sheet, expected_cols):
        for r in range(1, 100):
            row_vals = [cell.value for cell in sheet[r]]
            if not any(row_vals):
                continue
            cleaned_vals = [str(val).strip().lower() if val is not None else "" for val in row_vals]
            if "scenario_id" in cleaned_vals:
                return r, [str(val).strip() if val is not None else "" for val in row_vals]
        return None, None

    def get_val(row_vals, col_name, header_indices, row_num, sheet_name, required=True, val_type=str):
        idx = header_indices.get(col_name.lower())
        if idx is None or idx >= len(row_vals):
            if required:
                return None, f"{sheet_name}: Required column '{col_name}' is missing in sheet headers."
            return None, None
        val = row_vals[idx]
        if val is None or str(val).strip() == "":
            if required:
                return None, f"{sheet_name} Row {row_num}: Required field '{col_name}' is missing or empty."
            return None, None
        
        try:
            if val_type == int:
                return int(float(val)), None
            elif val_type == float:
                return float(val), None
            elif val_type == bool:
                s = str(val).strip().lower()
                if s in ("1", "true", "yes"):
                    return True, None
                elif s in ("0", "false", "no"):
                    return False, None
                else:
                    return None, f"{sheet_name} Row {row_num}: Field '{col_name}' must be a boolean (0 or 1, True or False). Got '{val}'."
            elif val_type == str:
                return str(val).strip(), None
            return val, None
        except (ValueError, TypeError):
            return None, f"{sheet_name} Row {row_num}: Field '{col_name}' has invalid format. Expected {val_type.__name__}, got '{val}'."

    try:
        sheet_names = wb.sheetnames
        sheet_map = {name.lower(): name for name in sheet_names}
        
        # Parsed entities in memory
        scenario_profiles_to_import = {}  # scenario_id -> ScenarioProfile
        appliances_to_import = []
        energy_assets_to_import = {}      # scenario_id -> EnergyAsset
        interval_inputs_to_import = []
        baseline_schedules_to_import = []
        
        # 1. Parse Scenario_Profiles
        sp_sheet_name = sheet_map.get("scenario_profiles")
        if not sp_sheet_name:
            errors.append("Sheet 'Scenario_Profiles' is missing from the workbook.")
        else:
            sheet = wb[sp_sheet_name]
            header_row, headers = find_sheet_header(sheet, ["scenario_id", "name", "building_type"])
            if not header_row:
                errors.append("Sheet 'Scenario_Profiles': Could not find a valid header row containing 'scenario_id'.")
            else:
                header_indices = {col.lower(): idx for idx, col in enumerate(headers) if col}
                for r in range(header_row + 1, sheet.max_row + 1):
                    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
                    if not any(v is not None for v in row_vals):
                        continue
                    
                    scenario_id, err = get_val(row_vals, "scenario_id", header_indices, r, "Scenario_Profiles", required=True, val_type=str)
                    if err: errors.append(err)
                    name, err = get_val(row_vals, "name", header_indices, r, "Scenario_Profiles", required=True, val_type=str)
                    if err: errors.append(err)
                    timezone, err = get_val(row_vals, "timezone", header_indices, r, "Scenario_Profiles", required=False, val_type=str)
                    if err: errors.append(err)
                    if not timezone: timezone = "Asia/Karachi"
                    
                    building_type, err = get_val(row_vals, "building_type", header_indices, r, "Scenario_Profiles", required=True, val_type=str)
                    if err: errors.append(err)
                    area_m2, err = get_val(row_vals, "area_m2", header_indices, r, "Scenario_Profiles", required=True, val_type=float)
                    if err: errors.append(err)
                    room_count, err = get_val(row_vals, "room_count", header_indices, r, "Scenario_Profiles", required=True, val_type=int)
                    if err: errors.append(err)
                    max_occupancy, err = get_val(row_vals, "max_occupancy", header_indices, r, "Scenario_Profiles", required=True, val_type=int)
                    if err: errors.append(err)
                    
                    insulation_level, err = get_val(row_vals, "insulation_level", header_indices, r, "Scenario_Profiles", required=True, val_type=str)
                    if err: errors.append(err)
                    elif insulation_level and insulation_level.lower() not in ("low", "medium", "high"):
                        errors.append(f"Scenario_Profiles Row {r}: insulation_level must be 'Low', 'Medium', or 'High' (got '{insulation_level}').")
                    
                    sun_exposure, err = get_val(row_vals, "sun_exposure", header_indices, r, "Scenario_Profiles", required=True, val_type=str)
                    if err: errors.append(err)
                    elif sun_exposure and sun_exposure.lower() not in ("low", "medium", "high"):
                        errors.append(f"Scenario_Profiles Row {r}: sun_exposure must be 'Low', 'Medium', or 'High' (got '{sun_exposure}').")
                    
                    comfort_min_c, err = get_val(row_vals, "comfort_min_c", header_indices, r, "Scenario_Profiles", required=True, val_type=float)
                    if err: errors.append(err)
                    comfort_max_c, err = get_val(row_vals, "comfort_max_c", header_indices, r, "Scenario_Profiles", required=True, val_type=float)
                    if err: errors.append(err)
                    
                    vulnerable_occupants, err = get_val(row_vals, "vulnerable_occupants", header_indices, r, "Scenario_Profiles", required=True, val_type=int)
                    if err: errors.append(err)
                    budget_pkr_per_day, err = get_val(row_vals, "budget_pkr_per_day", header_indices, r, "Scenario_Profiles", required=True, val_type=float)
                    if err: errors.append(err)
                    maximum_grid_demand_kw, err = get_val(row_vals, "maximum_grid_demand_kw", header_indices, r, "Scenario_Profiles", required=True, val_type=float)
                    if err: errors.append(err)
                    
                    # Range validations
                    if comfort_min_c is not None and comfort_max_c is not None and comfort_min_c >= comfort_max_c:
                        errors.append(f"Scenario_Profiles Row {r}: comfort_min_c ({comfort_min_c}) must be less than comfort_max_c ({comfort_max_c}).")
                    if area_m2 is not None and area_m2 <= 0:
                        errors.append(f"Scenario_Profiles Row {r}: area_m2 must be positive (got {area_m2}).")
                    if room_count is not None and room_count <= 0:
                        errors.append(f"Scenario_Profiles Row {r}: room_count must be positive (got {room_count}).")
                    if max_occupancy is not None and max_occupancy < 0:
                        errors.append(f"Scenario_Profiles Row {r}: max_occupancy must be non-negative (got {max_occupancy}).")
                    if vulnerable_occupants is not None and vulnerable_occupants < 0:
                        errors.append(f"Scenario_Profiles Row {r}: vulnerable_occupants must be non-negative (got {vulnerable_occupants}).")
                    if budget_pkr_per_day is not None and budget_pkr_per_day < 0:
                        warnings.append(f"Scenario_Profiles Row {r}: budget_pkr_per_day is negative or zero (got {budget_pkr_per_day}).")
                    if maximum_grid_demand_kw is not None and maximum_grid_demand_kw <= 0:
                        errors.append(f"Scenario_Profiles Row {r}: maximum_grid_demand_kw must be positive (got {maximum_grid_demand_kw}).")
                    
                    if not err and scenario_id:
                        scenario_profiles_to_import[scenario_id] = ScenarioProfile(
                            scenario_id=scenario_id,
                            name=name,
                            timezone=timezone,
                            building_type=building_type,
                            area_m2=area_m2,
                            room_count=room_count,
                            max_occupancy=max_occupancy,
                            insulation_level=insulation_level,
                            sun_exposure=sun_exposure,
                            comfort_min_c=comfort_min_c,
                            comfort_max_c=comfort_max_c,
                            vulnerable_occupants=vulnerable_occupants,
                            budget_pkr_per_day=budget_pkr_per_day,
                            maximum_grid_demand_kw=maximum_grid_demand_kw,
                        )

        # 2. Parse Appliances
        app_sheet_name = sheet_map.get("appliances")
        if not app_sheet_name:
            errors.append("Sheet 'Appliances' is missing from the workbook.")
        else:
            sheet = wb[app_sheet_name]
            header_row, headers = find_sheet_header(sheet, ["scenario_id", "appliance_id", "appliance_type"])
            if not header_row:
                errors.append("Sheet 'Appliances': Could not find a valid header row containing 'scenario_id'.")
            else:
                header_indices = {col.lower(): idx for idx, col in enumerate(headers) if col}
                for r in range(header_row + 1, sheet.max_row + 1):
                    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
                    if not any(v is not None for v in row_vals):
                        continue
                    
                    scenario_id, err = get_val(row_vals, "scenario_id", header_indices, r, "Appliances", required=True, val_type=str)
                    if err: errors.append(err)
                    appliance_id, err = get_val(row_vals, "appliance_id", header_indices, r, "Appliances", required=True, val_type=str)
                    if err: errors.append(err)
                    zone_id, err = get_val(row_vals, "zone_id", header_indices, r, "Appliances", required=True, val_type=str)
                    if err: errors.append(err)
                    appliance_type, err = get_val(row_vals, "appliance_type", header_indices, r, "Appliances", required=True, val_type=str)
                    if err: errors.append(err)
                    quantity, err = get_val(row_vals, "quantity", header_indices, r, "Appliances", required=True, val_type=int)
                    if err: errors.append(err)
                    rated_power_kw, err = get_val(row_vals, "rated_power_kw", header_indices, r, "Appliances", required=True, val_type=float)
                    if err: errors.append(err)
                    cooling_capacity_kw, err = get_val(row_vals, "cooling_capacity_kw", header_indices, r, "Appliances", required=True, val_type=float)
                    if err: errors.append(err)
                    efficiency_label, err = get_val(row_vals, "efficiency_label", header_indices, r, "Appliances", required=False, val_type=str)
                    if err: errors.append(err)
                    min_runtime_minutes, err = get_val(row_vals, "min_runtime_minutes", header_indices, r, "Appliances", required=True, val_type=int)
                    if err: errors.append(err)
                    min_setpoint_c, err = get_val(row_vals, "min_setpoint_c", header_indices, r, "Appliances", required=False, val_type=float)
                    if err: errors.append(err)
                    max_setpoint_c, err = get_val(row_vals, "max_setpoint_c", header_indices, r, "Appliances", required=False, val_type=float)
                    if err: errors.append(err)
                    
                    if appliance_type:
                        app_type_lower = appliance_type.lower()
                        if "inverter ac" in app_type_lower or "ac" in app_type_lower:
                            appliance_type = "Inverter AC"
                        elif "ceiling fan" in app_type_lower or "fan" in app_type_lower:
                            appliance_type = "Ceiling fan"
                        else:
                            errors.append(f"Appliances Row {r}: appliance_type must be 'Inverter AC' or 'Ceiling fan' (got '{appliance_type}').")
                    
                    if quantity is not None and quantity <= 0:
                        errors.append(f"Appliances Row {r}: quantity must be positive (got {quantity}).")
                    if rated_power_kw is not None and rated_power_kw <= 0:
                        errors.append(f"Appliances Row {r}: rated_power_kw must be positive (got {rated_power_kw}).")
                    if cooling_capacity_kw is not None and cooling_capacity_kw < 0:
                        errors.append(f"Appliances Row {r}: cooling_capacity_kw must be non-negative (got {cooling_capacity_kw}).")
                    if min_runtime_minutes is not None and min_runtime_minutes < 0:
                        errors.append(f"Appliances Row {r}: min_runtime_minutes must be non-negative (got {min_runtime_minutes}).")
                    if min_setpoint_c is not None and max_setpoint_c is not None:
                        if appliance_type == "Inverter AC" and min_setpoint_c >= max_setpoint_c:
                            errors.append(f"Appliances Row {r}: min_setpoint_c ({min_setpoint_c}) must be less than max_setpoint_c ({max_setpoint_c}).")
                    
                    if not err and appliance_id:
                        appliances_to_import.append(Appliance(
                            scenario_id=scenario_id,
                            appliance_id=appliance_id,
                            zone_id=zone_id,
                            appliance_type=appliance_type,
                            quantity=quantity,
                            rated_power_kw=rated_power_kw,
                            cooling_capacity_kw=cooling_capacity_kw,
                            efficiency_label=efficiency_label,
                            min_runtime_minutes=min_runtime_minutes,
                            min_setpoint_c=min_setpoint_c,
                            max_setpoint_c=max_setpoint_c,
                        ))

        # 3. Parse Energy_Assets
        ea_sheet_name = sheet_map.get("energy_assets")
        if not ea_sheet_name:
            errors.append("Sheet 'Energy_Assets' is missing from the workbook.")
        else:
            sheet = wb[ea_sheet_name]
            header_row, headers = find_sheet_header(sheet, ["scenario_id", "solar_capacity_kw", "battery_capacity_kwh"])
            if not header_row:
                errors.append("Sheet 'Energy_Assets': Could not find a valid header row containing 'scenario_id'.")
            else:
                header_indices = {col.lower(): idx for idx, col in enumerate(headers) if col}
                for r in range(header_row + 1, sheet.max_row + 1):
                    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
                    if not any(v is not None for v in row_vals):
                        continue
                    
                    scenario_id, err = get_val(row_vals, "scenario_id", header_indices, r, "Energy_Assets", required=True, val_type=str)
                    if err: errors.append(err)
                    solar_capacity_kw, err = get_val(row_vals, "solar_capacity_kw", header_indices, r, "Energy_Assets", required=True, val_type=float)
                    if err: errors.append(err)
                    solar_conversion_efficiency, err = get_val(row_vals, "solar_conversion_efficiency", header_indices, r, "Energy_Assets", required=True, val_type=float)
                    if err: errors.append(err)
                    battery_capacity_kwh, err = get_val(row_vals, "battery_capacity_kwh", header_indices, r, "Energy_Assets", required=True, val_type=float)
                    if err: errors.append(err)
                    initial_soc_kwh, err = get_val(row_vals, "initial_soc_kwh", header_indices, r, "Energy_Assets", required=True, val_type=float)
                    if err: errors.append(err)
                    minimum_reserve_kwh, err = get_val(row_vals, "minimum_reserve_kwh", header_indices, r, "Energy_Assets", required=True, val_type=float)
                    if err: errors.append(err)
                    max_charge_kw, err = get_val(row_vals, "max_charge_kw", header_indices, r, "Energy_Assets", required=True, val_type=float)
                    if err: errors.append(err)
                    max_discharge_kw, err = get_val(row_vals, "max_discharge_kw", header_indices, r, "Energy_Assets", required=True, val_type=float)
                    if err: errors.append(err)
                    charge_efficiency, err = get_val(row_vals, "charge_efficiency", header_indices, r, "Energy_Assets", required=True, val_type=float)
                    if err: errors.append(err)
                    discharge_efficiency, err = get_val(row_vals, "discharge_efficiency", header_indices, r, "Energy_Assets", required=True, val_type=float)
                    if err: errors.append(err)
                    
                    if solar_capacity_kw is not None and solar_capacity_kw < 0:
                        errors.append(f"Energy_Assets Row {r}: solar_capacity_kw cannot be negative (got {solar_capacity_kw}).")
                    if solar_conversion_efficiency is not None and not (0 <= solar_conversion_efficiency <= 1.0):
                        errors.append(f"Energy_Assets Row {r}: solar_conversion_efficiency must be between 0 and 1 (got {solar_conversion_efficiency}).")
                    if battery_capacity_kwh is not None and battery_capacity_kwh < 0:
                        errors.append(f"Energy_Assets Row {r}: battery_capacity_kwh cannot be negative (got {battery_capacity_kwh}).")
                    if battery_capacity_kwh is not None and battery_capacity_kwh > 0:
                        if initial_soc_kwh is not None and not (0 <= initial_soc_kwh <= battery_capacity_kwh):
                            errors.append(f"Energy_Assets Row {r}: initial_soc_kwh must be between 0 and battery capacity ({battery_capacity_kwh}) (got {initial_soc_kwh}).")
                        if minimum_reserve_kwh is not None and not (0 <= minimum_reserve_kwh <= battery_capacity_kwh):
                            errors.append(f"Energy_Assets Row {r}: minimum_reserve_kwh must be between 0 and battery capacity ({battery_capacity_kwh}) (got {minimum_reserve_kwh}).")
                    if max_charge_kw is not None and max_charge_kw < 0:
                        errors.append(f"Energy_Assets Row {r}: max_charge_kw cannot be negative (got {max_charge_kw}).")
                    if max_discharge_kw is not None and max_discharge_kw < 0:
                        errors.append(f"Energy_Assets Row {r}: max_discharge_kw cannot be negative (got {max_discharge_kw}).")
                    if charge_efficiency is not None and not (0 < charge_efficiency <= 1.0):
                        errors.append(f"Energy_Assets Row {r}: charge_efficiency must be between 0 and 1 (got {charge_efficiency}).")
                    if discharge_efficiency is not None and not (0 < discharge_efficiency <= 1.0):
                        errors.append(f"Energy_Assets Row {r}: discharge_efficiency must be between 0 and 1 (got {discharge_efficiency}).")
                    
                    if not err and scenario_id:
                        energy_assets_to_import[scenario_id] = EnergyAsset(
                            scenario_id=scenario_id,
                            solar_capacity_kw=solar_capacity_kw,
                            solar_conversion_efficiency=solar_conversion_efficiency,
                            battery_capacity_kwh=battery_capacity_kwh,
                            initial_soc_kwh=initial_soc_kwh,
                            minimum_reserve_kwh=minimum_reserve_kwh,
                            max_charge_kw=max_charge_kw,
                            max_discharge_kw=max_discharge_kw,
                            charge_efficiency=charge_efficiency,
                            discharge_efficiency=discharge_efficiency,
                        )

        # 4. Parse Interval_Inputs
        ii_sheet_name = sheet_map.get("interval_inputs")
        if not ii_sheet_name:
            errors.append("Sheet 'Interval_Inputs' is missing from the workbook.")
        else:
            sheet = wb[ii_sheet_name]
            header_row, headers = find_sheet_header(sheet, ["scenario_id", "timestamp_local", "temperature_c"])
            if not header_row:
                errors.append("Sheet 'Interval_Inputs': Could not find a valid header row containing 'scenario_id'.")
            else:
                header_indices = {col.lower(): idx for idx, col in enumerate(headers) if col}
                scenario_timestamps = {}
                
                for r in range(header_row + 1, sheet.max_row + 1):
                    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
                    if not any(v is not None for v in row_vals):
                        continue
                    
                    scenario_id, err = get_val(row_vals, "scenario_id", header_indices, r, "Interval_Inputs", required=True, val_type=str)
                    if err: errors.append(err)
                    
                    timestamp_local = None
                    idx = header_indices.get("timestamp_local")
                    if idx is not None and idx < len(row_vals):
                        timestamp_local = row_vals[idx]
                    
                    dt_val = None
                    if timestamp_local is None or str(timestamp_local).strip() == "":
                        errors.append(f"Interval_Inputs Row {r}: Required field 'timestamp_local' is missing or empty.")
                    else:
                        if isinstance(timestamp_local, datetime):
                            dt_val = timestamp_local
                        else:
                            ts_str = str(timestamp_local).strip()
                            try:
                                if "T" in ts_str:
                                    dt_val = datetime.fromisoformat(ts_str)
                                else:
                                    dt_val = datetime.strptime(ts_str, "%Y-%m-%d %H:%M:%S")
                            except ValueError:
                                errors.append(f"Interval_Inputs Row {r}: Invalid timestamp format '{ts_str}'. Must be ISO 8601.")
                    
                    interval_minutes, err = get_val(row_vals, "interval_minutes", header_indices, r, "Interval_Inputs", required=True, val_type=int)
                    if err: errors.append(err)
                    temperature_c, err = get_val(row_vals, "temperature_c", header_indices, r, "Interval_Inputs", required=True, val_type=float)
                    if err: errors.append(err)
                    relative_humidity_pct, err = get_val(row_vals, "relative_humidity_pct", header_indices, r, "Interval_Inputs", required=True, val_type=float)
                    if err: errors.append(err)
                    
                    heat_index_c, err = get_val(row_vals, "heat_index_c", header_indices, r, "Interval_Inputs", required=False, val_type=float)
                    if err: errors.append(err)
                    if heat_index_c is None and temperature_c is not None:
                        heat_index_c = temperature_c
                        
                    solar_irradiance_w_m2, err = get_val(row_vals, "solar_irradiance_w_m2", header_indices, r, "Interval_Inputs", required=True, val_type=float)
                    if err: errors.append(err)
                    solar_available_kw, err = get_val(row_vals, "solar_available_kw", header_indices, r, "Interval_Inputs", required=True, val_type=float)
                    if err: errors.append(err)
                    occupancy_count, err = get_val(row_vals, "occupancy_count", header_indices, r, "Interval_Inputs", required=True, val_type=int)
                    if err: errors.append(err)
                    grid_available, err = get_val(row_vals, "grid_available", header_indices, r, "Interval_Inputs", required=True, val_type=bool)
                    if err: errors.append(err)
                    
                    tariff_type, err = get_val(row_vals, "tariff_type", header_indices, r, "Interval_Inputs", required=True, val_type=str)
                    if err: errors.append(err)
                    elif tariff_type:
                        tariff_type_up = tariff_type.upper()
                        if tariff_type_up not in ("PEAK", "OFF_PEAK"):
                            errors.append(f"Interval_Inputs Row {r}: tariff_type must be 'PEAK' or 'OFF_PEAK' (got '{tariff_type}').")
                        else:
                            tariff_type = tariff_type_up
                            
                    tariff_pkr_per_kwh, err = get_val(row_vals, "tariff_pkr_per_kwh", header_indices, r, "Interval_Inputs", required=True, val_type=float)
                    if err: errors.append(err)
                    grid_carbon_kgco2_per_kwh, err = get_val(row_vals, "grid_carbon_kgco2_per_kwh", header_indices, r, "Interval_Inputs", required=True, val_type=float)
                    if err: errors.append(err)
                    non_cooling_load_kw, err = get_val(row_vals, "non_cooling_load_kw", header_indices, r, "Interval_Inputs", required=True, val_type=float)
                    if err: errors.append(err)
                    source_missing_flag, err = get_val(row_vals, "source_missing_flag", header_indices, r, "Interval_Inputs", required=True, val_type=int)
                    if err: errors.append(err)
                    
                    if temperature_c is not None and not (0 <= temperature_c <= 60.0):
                        errors.append(f"Interval_Inputs Row {r}: temperature_c ({temperature_c}) is outside reasonable range [0, 60].")
                    if relative_humidity_pct is not None and not (0 <= relative_humidity_pct <= 100.0):
                        errors.append(f"Interval_Inputs Row {r}: relative_humidity_pct ({relative_humidity_pct}) is outside reasonable range [0, 100].")
                    if solar_irradiance_w_m2 is not None and solar_irradiance_w_m2 < 0:
                        errors.append(f"Interval_Inputs Row {r}: solar_irradiance_w_m2 cannot be negative (got {solar_irradiance_w_m2}).")
                    if solar_available_kw is not None and solar_available_kw < 0:
                        errors.append(f"Interval_Inputs Row {r}: solar_available_kw cannot be negative (got {solar_available_kw}).")
                    if occupancy_count is not None and occupancy_count < 0:
                        errors.append(f"Interval_Inputs Row {r}: occupancy_count cannot be negative (got {occupancy_count}).")
                    if tariff_pkr_per_kwh is not None and tariff_pkr_per_kwh < 0:
                        errors.append(f"Interval_Inputs Row {r}: tariff_pkr_per_kwh cannot be negative (got {tariff_pkr_per_kwh}).")
                    if non_cooling_load_kw is not None and non_cooling_load_kw < 0:
                        errors.append(f"Interval_Inputs Row {r}: non_cooling_load_kw cannot be negative (got {non_cooling_load_kw}).")
                    if source_missing_flag is not None and source_missing_flag not in (0, 1):
                        errors.append(f"Interval_Inputs Row {r}: source_missing_flag must be 0 or 1 (got {source_missing_flag}).")
                    
                    if dt_val and scenario_id:
                        scenario_timestamps.setdefault(scenario_id, []).append((dt_val, r, str(timestamp_local)))
                        
                    if not err and dt_val and scenario_id:
                        interval_inputs_to_import.append(IntervalInput(
                            scenario_id=scenario_id,
                            timestamp_local=dt_val,
                            interval_minutes=interval_minutes,
                            temperature_c=temperature_c,
                            relative_humidity_pct=relative_humidity_pct,
                            heat_index_c=heat_index_c,
                            solar_irradiance_w_m2=solar_irradiance_w_m2,
                            solar_available_kw=solar_available_kw,
                            occupancy_count=occupancy_count,
                            grid_available=grid_available,
                            tariff_type=tariff_type,
                            tariff_pkr_per_kwh=tariff_pkr_per_kwh,
                            grid_carbon_kgco2_per_kwh=grid_carbon_kgco2_per_kwh,
                            non_cooling_load_kw=non_cooling_load_kw,
                            source_missing_flag=source_missing_flag,
                        ))

                for s_id, tlist in scenario_timestamps.items():
                    is_sorted = True
                    for i in range(1, len(tlist)):
                        prev_dt, prev_r, prev_val = tlist[i-1]
                        curr_dt, curr_r, curr_val = tlist[i]
                        if curr_dt < prev_dt:
                            errors.append(f"Interval_Inputs (Scenario {s_id}): Timestamps are not sorted. Row {curr_r} ({curr_val}) is before Row {prev_r} ({prev_val}).")
                            is_sorted = False
                            break
                    
                    if is_sorted and len(tlist) > 1:
                        gaps = 0
                        for i in range(1, len(tlist)):
                            prev_dt, prev_r, _ = tlist[i-1]
                            curr_dt, curr_r, _ = tlist[i]
                            diff = (curr_dt - prev_dt).total_seconds() / 60.0
                            if diff != 15.0:
                                gaps += 1
                                if gaps <= 5:
                                    errors.append(f"Interval_Inputs (Scenario {s_id}) Row {curr_r}: Timestamp gap detected. Interval from row {prev_r} to row {curr_r} is {diff} minutes (expected 15).")
                        if gaps > 5:
                            errors.append(f"Interval_Inputs (Scenario {s_id}): Total of {gaps} timestamp gaps detected.")

        # 5. Parse Baseline_Schedule
        bs_sheet_name = sheet_map.get("baseline_schedule") or sheet_map.get("baseline_schedules")
        if not bs_sheet_name:
            errors.append("Sheet 'Baseline_Schedule' is missing from the workbook.")
        else:
            sheet = wb[bs_sheet_name]
            header_row, headers = find_sheet_header(sheet, ["scenario_id", "timestamp_local", "baseline_ac_units_on"])
            if not header_row:
                errors.append("Sheet 'Baseline_Schedule': Could not find a valid header row containing 'scenario_id'.")
            else:
                header_indices = {col.lower(): idx for idx, col in enumerate(headers) if col}
                scenario_bs_timestamps = {}
                
                for r in range(header_row + 1, sheet.max_row + 1):
                    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
                    if not any(v is not None for v in row_vals):
                        continue
                    
                    scenario_id, err = get_val(row_vals, "scenario_id", header_indices, r, "Baseline_Schedule", required=True, val_type=str)
                    if err: errors.append(err)
                    
                    timestamp_local = None
                    idx = header_indices.get("timestamp_local")
                    if idx is not None and idx < len(row_vals):
                        timestamp_local = row_vals[idx]
                    
                    dt_val = None
                    if timestamp_local is None or str(timestamp_local).strip() == "":
                        errors.append(f"Baseline_Schedule Row {r}: Required field 'timestamp_local' is missing or empty.")
                    else:
                        if isinstance(timestamp_local, datetime):
                            dt_val = timestamp_local
                        else:
                            ts_str = str(timestamp_local).strip()
                            try:
                                if "T" in ts_str:
                                    dt_val = datetime.fromisoformat(ts_str)
                                else:
                                    dt_val = datetime.strptime(ts_str, "%Y-%m-%d %H:%M:%S")
                            except ValueError:
                                errors.append(f"Baseline_Schedule Row {r}: Invalid timestamp format '{ts_str}'. Must be ISO 8601.")
                    
                    baseline_ac_units_on, err = get_val(row_vals, "baseline_ac_units_on", header_indices, r, "Baseline_Schedule", required=True, val_type=int)
                    if err: errors.append(err)
                    baseline_ac_setpoint_c, err = get_val(row_vals, "baseline_ac_setpoint_c", header_indices, r, "Baseline_Schedule", required=False, val_type=float)
                    if err: errors.append(err)
                    baseline_fan_units_on, err = get_val(row_vals, "baseline_fan_units_on", header_indices, r, "Baseline_Schedule", required=True, val_type=int)
                    if err: errors.append(err)
                    baseline_other_cooling_kw, err = get_val(row_vals, "baseline_other_cooling_kw", header_indices, r, "Baseline_Schedule", required=True, val_type=float)
                    if err: errors.append(err)
                    baseline_rule, err = get_val(row_vals, "baseline_rule", header_indices, r, "Baseline_Schedule", required=False, val_type=str)
                    if err: errors.append(err)
                    
                    if baseline_ac_units_on is not None and baseline_ac_units_on < 0:
                        errors.append(f"Baseline_Schedule Row {r}: baseline_ac_units_on cannot be negative (got {baseline_ac_units_on}).")
                    if baseline_fan_units_on is not None and baseline_fan_units_on < 0:
                        errors.append(f"Baseline_Schedule Row {r}: baseline_fan_units_on cannot be negative (got {baseline_fan_units_on}).")
                    if baseline_other_cooling_kw is not None and baseline_other_cooling_kw < 0:
                        errors.append(f"Baseline_Schedule Row {r}: baseline_other_cooling_kw cannot be negative (got {baseline_other_cooling_kw}).")
                    
                    if dt_val and scenario_id:
                        scenario_bs_timestamps.setdefault(scenario_id, []).append((dt_val, r, str(timestamp_local)))
                        
                    if not err and dt_val and scenario_id:
                        baseline_schedules_to_import.append(BaselineSchedule(
                            scenario_id=scenario_id,
                            timestamp_local=dt_val,
                            baseline_ac_units_on=baseline_ac_units_on,
                            baseline_ac_setpoint_c=baseline_ac_setpoint_c,
                            baseline_fan_units_on=baseline_fan_units_on,
                            baseline_other_cooling_kw=baseline_other_cooling_kw,
                            baseline_rule=baseline_rule,
                        ))

                for s_id, tlist in scenario_bs_timestamps.items():
                    is_sorted = True
                    for i in range(1, len(tlist)):
                        prev_dt, prev_r, prev_val = tlist[i-1]
                        curr_dt, curr_r, curr_val = tlist[i]
                        if curr_dt < prev_dt:
                            errors.append(f"Baseline_Schedule (Scenario {s_id}): Timestamps are not sorted. Row {curr_r} ({curr_val}) is before Row {prev_r} ({prev_val}).")
                            is_sorted = False
                            break
                    
                    if is_sorted and len(tlist) > 1:
                        gaps = 0
                        for i in range(1, len(tlist)):
                            prev_dt, prev_r, _ = tlist[i-1]
                            curr_dt, curr_r, _ = tlist[i]
                            diff = (curr_dt - prev_dt).total_seconds() / 60.0
                            if diff != 15.0:
                                gaps += 1
                                if gaps <= 5:
                                    errors.append(f"Baseline_Schedule (Scenario {s_id}) Row {curr_r}: Timestamp gap detected. Interval from row {prev_r} to row {curr_r} is {diff} minutes (expected 15).")
                        if gaps > 5:
                            errors.append(f"Baseline_Schedule (Scenario {s_id}): Total of {gaps} timestamp gaps detected.")

        # Database Insertion (Only if no errors are present)
        if len(errors) == 0:
            scenarios_seen = set(scenario_profiles_to_import.keys()) | set(scenario_timestamps.keys()) | set(scenario_bs_timestamps.keys())
            
            for s_id in scenarios_seen:
                # Cascade deletes using SQLAlchemy session to be clean
                db.query(Appliance).filter_by(scenario_id=s_id).delete()
                db.query(IntervalInput).filter_by(scenario_id=s_id).delete()
                db.query(BaselineSchedule).filter_by(scenario_id=s_id).delete()
                db.query(EnergyAsset).filter_by(scenario_id=s_id).delete()
                db.query(ScenarioProfile).filter_by(scenario_id=s_id).delete()
            
            for sp in scenario_profiles_to_import.values():
                db.add(sp)
            db.bulk_save_objects(appliances_to_import)
            for ea in energy_assets_to_import.values():
                db.add(ea)
            db.bulk_save_objects(interval_inputs_to_import)
            db.bulk_save_objects(baseline_schedules_to_import)
            
            db.commit()
            
            imported_counts = {
                "scenario_profiles": len(scenario_profiles_to_import),
                "appliances": len(appliances_to_import),
                "energy_assets": len(energy_assets_to_import),
                "interval_inputs": len(interval_inputs_to_import),
                "baseline_schedules": len(baseline_schedules_to_import),
            }
            
    except Exception as e:
        db.rollback()
        errors.append(f"Import process failed: {str(e)}")
    finally:
        if wb:
            try:
                wb.close()
            except Exception:
                pass
                
    success = len(errors) == 0
    return {
        "success": success,
        "errors": errors,
        "warnings": warnings,
        "imported_counts": imported_counts,
    }

